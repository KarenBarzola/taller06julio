from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from eventos.forms import ClienteFormulario
from eventos.models import Cliente

# Create your views here.
'''def bienvenida(request):
    pagina = loader.get_template('')
    return HttpResponse(pagina.render())
'''
def agregar_cliente(request):
    pagina = loader.get_template('agregar_clientes.html')
    if request.method == 'GET':
        formulario = ClienteFormulario

    elif request.method == 'POST':
        formulario = ClienteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_cliente(request, idCliente):
    pagina = loader.get_template('ver_cliente.html')
    #cliente = Cliente.objects.get(pk=idCliente)
    cliente = get_object_or_404(Cliente, pk=idCliente)
    mensaje = {'cliente': cliente}
    return HttpResponse(pagina.render(mensaje, request))

def editar_cliente(request, idCliente):
    pagina = loader.get_template('editar_cliente.html')
    cliente = get_object_or_404(Cliente, pk=idCliente)
    if request.method == 'GET':
        formulario = ClienteFormulario(instance=cliente)
    elif request.method == 'POST':
        formulario = ClienteFormulario(request.POST, instance=cliente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    mensaje = {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))

def eliminar_cliente(request, idCliente):
    cliente = get_object_or_404(Cliente, pk=idCliente)
    if cliente:
        cliente.delete()
        return redirect('inicio')

def generar_archivo(request):
    #clientes = Cliente.objects.all()
    clientes = Cliente.objects.order_by('apellido')

    wb = Workbook()

    ws = wb.active

    ws['B1'] = 'REPORTE CLIENTES'

    ws.merge_cells('B1:E1')

    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'PAGO'
    ws['E3'] = 'DIRECCION'
    ws['F3'] = 'CORREO'
    cont = 5

    for cliente in clientes:
        ws.cell(row=cont, column=2).value = cliente.nombre
        ws.cell(row=cont, column=3).value = cliente.apellido
        ws.cell(row=cont, column=4).value = cliente.pago
        ws.cell(row=cont, column=5).value = cliente.direccion
        ws.cell(row=cont, column=6).value = cliente.correo

        cont = cont + 1

    nombre_archivo = "ReporteClientesExcel.xlsx"
    response = HttpResponse(content_type="applications/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

